import openpyxl
# import Xlsxutils

# file-workbook-sheets-rows-cells
file=r"C:\Users\Chandan\Documents\Test.xlsx"
# print(Xlsxutils.getColumeCount(file,"Sheet1"))
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]

rows=sheet.max_row
cols=sheet.max_column

for i in range(1,rows):
    for j in range(1,cols):
        print(sheet.cell(i,j).value)
# data=200
# sheet.cell(2,5).value=data
# workbook.save(file)